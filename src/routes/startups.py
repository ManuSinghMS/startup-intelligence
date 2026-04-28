"""
Startup CRUD routes — with tag management, deactivation, and CSV import.
"""
import uuid
import csv
import io
from datetime import datetime
from typing import Optional, List

from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from pydantic import BaseModel

from src.db.database import get_db

router = APIRouter(prefix="/api/startups", tags=["startups"])


class StartupCreate(BaseModel):
    name: str
    legal_name: Optional[str] = None
    website: Optional[str] = None
    contact_email: Optional[str] = None
    contact_name: Optional[str] = None
    founder_name: Optional[str] = None
    cofounder_name: Optional[str] = None
    founder_linkedin_url: Optional[str] = None
    cofounder_linkedin_url: Optional[str] = None
    description: Optional[str] = None
    industry: Optional[str] = None
    secondary_industry: Optional[str] = None
    linkedin_url: Optional[str] = None
    twitter_handle: Optional[str] = None
    instagram_handle: Optional[str] = None
    stage: Optional[str] = None
    status: Optional[str] = None
    program_stream: Optional[str] = None
    tag: Optional[str] = "active"


class StartupUpdate(BaseModel):
    name: Optional[str] = None
    legal_name: Optional[str] = None
    website: Optional[str] = None
    contact_email: Optional[str] = None
    contact_name: Optional[str] = None
    founder_name: Optional[str] = None
    cofounder_name: Optional[str] = None
    founder_linkedin_url: Optional[str] = None
    cofounder_linkedin_url: Optional[str] = None
    description: Optional[str] = None
    industry: Optional[str] = None
    secondary_industry: Optional[str] = None
    linkedin_url: Optional[str] = None
    twitter_handle: Optional[str] = None
    instagram_handle: Optional[str] = None
    stage: Optional[str] = None
    status: Optional[str] = None
    program_stream: Optional[str] = None
    tag: Optional[str] = None


class BulkTagRequest(BaseModel):
    startup_ids: List[str]
    tag: str


@router.get("")
def list_startups(
    search: Optional[str] = Query(None, description="Search by name"),
    industry: Optional[str] = Query(None),
    tag: Optional[str] = Query(None, description="Filter by tag: active, alumni, not_active"),
    include_inactive: bool = Query(False, description="Include not_active companies"),
    limit: int = Query(100, le=500),
    offset: int = Query(0, ge=0)
):
    """List all startups with optional filters. Excludes not_active by default."""
    db = get_db()
    conditions = []
    params = []

    if search:
        conditions.append("(name LIKE ? OR legal_name LIKE ?)")
        params.extend([f"%{search}%", f"%{search}%"])
    if industry:
        conditions.append("(industry LIKE ? OR secondary_industry LIKE ?)")
        params.extend([f"%{industry}%", f"%{industry}%"])
    if tag:
        conditions.append("tag = ?")
        params.append(tag)
    elif not include_inactive:
        conditions.append("(tag IS NULL OR tag != 'not_active')")

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    # Get total count
    count = db.execute(f"SELECT COUNT(*) as c FROM startups {where}", params).fetchone()["c"]

    # Get startups with content count
    startups = db.execute(f"""
        SELECT s.*,
            (SELECT COUNT(*) FROM content_items ci WHERE ci.startup_id = s.id) as content_count,
            (SELECT MAX(ci.published_at) FROM content_items ci WHERE ci.startup_id = s.id) as last_activity
        FROM startups s
        {where}
        ORDER BY s.name
        LIMIT ? OFFSET ?
    """, params + [limit, offset]).fetchall()

    return {
        "total": count,
        "startups": [dict(s) for s in startups]
    }


@router.get("/{startup_id}")
def get_startup(startup_id: str):
    """Get a single startup with its recent content."""
    db = get_db()
    startup = db.execute(
        "SELECT * FROM startups WHERE id = ?", (startup_id,)
    ).fetchone()

    if not startup:
        raise HTTPException(status_code=404, detail="Startup not found")

    # Get recent content
    content = db.execute("""
        SELECT id, source_type, source_name, url, title, published_at,
               summary, classification, sentiment, created_at, ingestion_status
        FROM content_items
        WHERE startup_id = ? AND is_relevant = 1
        ORDER BY published_at DESC
        LIMIT 50
    """, (startup_id,)).fetchall()

    # Get content stats
    stats = db.execute("""
        SELECT classification, COUNT(*) as count
        FROM content_items
        WHERE startup_id = ? AND is_relevant = 1
        GROUP BY classification
    """, (startup_id,)).fetchall()

    return {
        "startup": dict(startup),
        "content": [dict(c) for c in content],
        "stats": {s["classification"]: s["count"] for s in stats}
    }


@router.post("", status_code=201)
def create_startup(data: StartupCreate):
    """Add a new startup."""
    db = get_db()
    startup_id = str(uuid.uuid4())
    now = datetime.utcnow().isoformat()

    db.execute(
        """INSERT INTO startups
        (id, name, legal_name, website, contact_email, contact_name,
         founder_name, cofounder_name, founder_linkedin_url, cofounder_linkedin_url,
         description, industry, secondary_industry, linkedin_url,
         twitter_handle, instagram_handle, stage, status, program_stream,
         tag, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (startup_id, data.name, data.legal_name, data.website,
         data.contact_email, data.contact_name,
         data.founder_name, data.cofounder_name,
         data.founder_linkedin_url, data.cofounder_linkedin_url,
         data.description,
         data.industry, data.secondary_industry, data.linkedin_url,
         data.twitter_handle, data.instagram_handle, data.stage,
         data.status, data.program_stream, data.tag or "active", now, now)
    )
    db.commit()

    return {"id": startup_id, "name": data.name}


@router.put("/{startup_id}")
def update_startup(startup_id: str, data: StartupUpdate):
    """Update a startup."""
    db = get_db()
    existing = db.execute(
        "SELECT * FROM startups WHERE id = ?", (startup_id,)
    ).fetchone()

    if not existing:
        raise HTTPException(status_code=404, detail="Startup not found")

    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")

    updates["updated_at"] = datetime.utcnow().isoformat()
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [startup_id]

    db.execute(f"UPDATE startups SET {set_clause} WHERE id = ?", values)
    db.commit()

    return {"id": startup_id, "updated": list(updates.keys())}


@router.delete("/{startup_id}")
def delete_startup(startup_id: str):
    """Delete a startup."""
    db = get_db()
    existing = db.execute(
        "SELECT id FROM startups WHERE id = ?", (startup_id,)
    ).fetchone()

    if not existing:
        raise HTTPException(status_code=404, detail="Startup not found")

    db.execute("DELETE FROM content_items WHERE startup_id = ?", (startup_id,))
    db.execute("DELETE FROM startups WHERE id = ?", (startup_id,))
    db.commit()

    return {"deleted": startup_id}


# --------------- Tag management ---------------

@router.patch("/{startup_id}/tag")
def update_tag(startup_id: str, tag: str = Query(..., description="Tag: active, alumni, not_active")):
    """Quick-update a startup's tag."""
    if tag not in ("active", "alumni", "not_active"):
        raise HTTPException(status_code=400, detail="Tag must be active, alumni, or not_active")

    db = get_db()
    existing = db.execute("SELECT id FROM startups WHERE id = ?", (startup_id,)).fetchone()
    if not existing:
        raise HTTPException(status_code=404, detail="Startup not found")

    db.execute(
        "UPDATE startups SET tag = ?, updated_at = ? WHERE id = ?",
        (tag, datetime.utcnow().isoformat(), startup_id)
    )
    db.commit()
    return {"id": startup_id, "tag": tag}


@router.put("/bulk-tag")
def bulk_tag(data: BulkTagRequest):
    """Set the same tag on multiple startups at once."""
    if data.tag not in ("active", "alumni", "not_active"):
        raise HTTPException(status_code=400, detail="Tag must be active, alumni, or not_active")

    db = get_db()
    now = datetime.utcnow().isoformat()
    updated = 0
    for sid in data.startup_ids:
        res = db.execute(
            "UPDATE startups SET tag = ?, updated_at = ? WHERE id = ?",
            (data.tag, now, sid)
        )
        updated += res.rowcount
    db.commit()
    return {"updated": updated, "tag": data.tag}


# --------------- Deactivation ---------------

@router.post("/{startup_id}/deactivate")
def deactivate_startup(startup_id: str):
    """Mark a startup as not_active."""
    db = get_db()
    existing = db.execute("SELECT id FROM startups WHERE id = ?", (startup_id,)).fetchone()
    if not existing:
        raise HTTPException(status_code=404, detail="Startup not found")

    db.execute(
        "UPDATE startups SET tag = 'not_active', updated_at = ? WHERE id = ?",
        (datetime.utcnow().isoformat(), startup_id)
    )
    db.commit()
    return {"id": startup_id, "tag": "not_active"}


@router.delete("/deactivated")
def remove_deactivated():
    """Delete ALL startups tagged not_active and their content."""
    db = get_db()
    deactivated = db.execute(
        "SELECT id, name FROM startups WHERE tag = 'not_active'"
    ).fetchall()

    if not deactivated:
        return {"deleted": 0, "companies": []}

    names = []
    for row in deactivated:
        db.execute("DELETE FROM content_items WHERE startup_id = ?", (row["id"],))
        db.execute("DELETE FROM summaries WHERE startup_id = ?", (row["id"],))
        db.execute("DELETE FROM startup_sources WHERE startup_id = ?", (row["id"],))
        db.execute("DELETE FROM startups WHERE id = ?", (row["id"],))
        names.append(row["name"])

    db.commit()
    return {"deleted": len(names), "companies": names}


# --------------- CSV / Excel Import ---------------

def _normalize(val):
    """Strip and return string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


@router.post("/import")
async def import_companies(file: UploadFile = File(...)):
    """
    Import companies from a CSV or Excel (.xlsx) file.
    Deduplicates by name (case-insensitive).
    """
    db = get_db()
    filename = file.filename.lower()
    content = await file.read()

    rows = []

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        # Row 3 has headers in Monday.com export, data starts at row 4
        # Column mapping: C1=Name, C3=Legal Name, C4=Stage, C5=Status,
        # C6=Email, C7=Contact, C33=Program Stream, C44=Industry, C45=Secondary
        for row_idx in range(4, ws.max_row + 1):
            name = ws.cell(row=row_idx, column=1).value
            if not name or str(name).strip() == "Subitems":
                continue
            rows.append({
                "name": str(name).strip(),
                "legal_name": _normalize(ws.cell(row=row_idx, column=3).value),
                "stage": _normalize(ws.cell(row=row_idx, column=4).value),
                "status": _normalize(ws.cell(row=row_idx, column=5).value),
                "contact_email": _normalize(ws.cell(row=row_idx, column=6).value),
                "contact_name": _normalize(ws.cell(row=row_idx, column=7).value),
                "program_stream": _normalize(ws.cell(row=row_idx, column=33).value),
                "industry": _normalize(ws.cell(row=row_idx, column=44).value),
                "secondary_industry": _normalize(ws.cell(row=row_idx, column=45).value),
            })

    elif filename.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for row_data in reader:
            name = (row_data.get("Name") or row_data.get("name") or "").strip()
            if not name or name == "Subitems":
                continue
            rows.append({
                "name": name,
                "legal_name": _normalize(row_data.get("Legal Name") or row_data.get("legal_name")),
                "stage": _normalize(row_data.get("Stage") or row_data.get("stage")),
                "status": _normalize(row_data.get("Status") or row_data.get("status")),
                "contact_email": _normalize(row_data.get("Email") or row_data.get("contact_email")),
                "contact_name": _normalize(row_data.get("Contact") or row_data.get("contact_name")),
                "program_stream": _normalize(row_data.get("Program Stream") or row_data.get("program_stream")),
                "industry": _normalize(row_data.get("Industry") or row_data.get("industry")),
                "secondary_industry": _normalize(row_data.get("Secondary Industry") or row_data.get("secondary_industry")),
            })
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Use .csv or .xlsx")

    imported = 0
    skipped = 0
    now = datetime.utcnow().isoformat()

    for row_data in rows:
        existing = db.execute(
            "SELECT id FROM startups WHERE LOWER(name) = LOWER(?)",
            (row_data["name"],)
        ).fetchone()

        if existing:
            skipped += 1
            continue

        startup_id = str(uuid.uuid4())
        db.execute(
            """INSERT INTO startups
            (id, name, legal_name, contact_email, contact_name, industry,
             secondary_industry, stage, status, program_stream, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (startup_id, row_data["name"], row_data["legal_name"],
             row_data["contact_email"], row_data["contact_name"],
             row_data["industry"], row_data["secondary_industry"],
             row_data["stage"], row_data["status"],
             row_data["program_stream"], now, now)
        )
        imported += 1

    db.commit()
    return {
        "imported": imported,
        "skipped": skipped,
        "total_in_file": len(rows),
    }


# --------------- Monday.com Sync ---------------

@router.post("/sync-monday")
async def sync_monday(dry_run: bool = Query(False, description="Preview only, don't write")):
    """
    Sync startups from Monday.com board.
    Requires MONDAY_API_TOKEN and MONDAY_BOARD_ID in .env.
    """
    from src.ingestion.monday_sync import is_configured, sync_from_monday
    if not is_configured():
        raise HTTPException(
            status_code=400,
            detail="Monday.com not configured. Set MONDAY_API_TOKEN and MONDAY_BOARD_ID in .env"
        )
    return await sync_from_monday(dry_run=dry_run)


@router.get("/monday-columns")
async def monday_columns():
    """Preview Monday.com board columns and current mapping."""
    from src.ingestion.monday_sync import is_configured, preview_columns
    if not is_configured():
        raise HTTPException(
            status_code=400,
            detail="Monday.com not configured. Set MONDAY_API_TOKEN and MONDAY_BOARD_ID in .env"
        )
    return await preview_columns()
