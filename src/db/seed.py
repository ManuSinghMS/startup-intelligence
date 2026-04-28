"""
Seed the database with companies from the Monday.com Excel export
and priority news/newsletter sources.
"""
import uuid
import sys
import os

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from dotenv import load_dotenv
load_dotenv()

from src.db.database import get_db, init_db


def import_monday_excel(filepath: str):
    """Import companies from the Monday.com board Excel export."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Row 3 has headers, data starts at row 4
    # Key columns from the Monday board:
    # C1: Name, C3: Legal Name, C4: Onboarding Stage, C5: Status,
    # C6: Contact Email, C7: Main contact name, C33: Program Stream
    # C44: Primary industry, C45: Secondary industry

    db = get_db()
    imported = 0
    skipped = 0

    for row in range(4, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name or name == "Subitems":
            continue

        legal_name = ws.cell(row=row, column=3).value
        stage = ws.cell(row=row, column=4).value
        status = ws.cell(row=row, column=5).value
        email = ws.cell(row=row, column=6).value
        contact = ws.cell(row=row, column=7).value
        program_stream = ws.cell(row=row, column=33).value
        industry = ws.cell(row=row, column=44).value
        secondary_industry = ws.cell(row=row, column=45).value

        # Check if startup already exists
        existing = db.execute(
            "SELECT id FROM startups WHERE name = ?", (name,)
        ).fetchone()

        if existing:
            skipped += 1
            continue

        startup_id = str(uuid.uuid4())
        db.execute(
            """INSERT INTO startups
            (id, name, legal_name, contact_email, contact_name, industry,
             secondary_industry, stage, status, program_stream)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (startup_id, name, legal_name, email, contact, industry,
             secondary_industry, stage, status, program_stream)
        )
        imported += 1

    db.commit()
    print(f"Imported {imported} startups from Monday.com ({skipped} skipped as duplicates)")


def seed_sources():
    """Seed priority news and newsletter sources."""
    db = get_db()

    sources = [
        # News outlets
        ("TechCrunch", "https://techcrunch.com", "https://techcrunch.com/feed/", "news", 5),
        ("VentureBeat", "https://venturebeat.com", "https://venturebeat.com/feed/", "news", 4),
        ("Crunchbase News", "https://news.crunchbase.com", "https://news.crunchbase.com/feed/", "news", 5),
        ("The Verge", "https://theverge.com", "https://www.theverge.com/rss/index.xml", "news", 3),
        ("Ars Technica", "https://arstechnica.com", "https://feeds.arstechnica.com/arstechnica/index", "news", 3),
        ("Wired", "https://wired.com", "https://www.wired.com/feed/rss", "news", 3),
        ("TechRadar", "https://techradar.com", "https://www.techradar.com/rss", "news", 2),
        ("Hacker News", "https://news.ycombinator.com", "https://hnrss.org/frontpage", "news", 4),
        ("Product Hunt", "https://producthunt.com", "https://www.producthunt.com/feed", "news", 3),

        # Canadian tech news (relevant for incubator)
        ("BetaKit", "https://betakit.com", "https://betakit.com/feed/", "news", 5),
        ("MaRS Discovery District", "https://marsdd.com", "https://marsdd.com/feed/", "news", 4),
        ("Communitech News", "https://communitech.ca", "https://www.communitech.ca/feed/", "news", 4),

        # Startup newsletters
        ("The Hustle", "https://thehustle.co", None, "newsletter", 3),
        ("StrictlyVC", "https://www.strictlyvc.com", None, "newsletter", 4),
        ("Mattermark Daily", "https://mattermark.com", None, "newsletter", 3),
        ("CB Insights", "https://www.cbinsights.com", "https://www.cbinsights.com/rss", "newsletter", 4),
        ("Inside.com Startups", "https://inside.com/startups", None, "newsletter", 3),

        # Press release sources
        ("PR Newswire", "https://prnewswire.com", "https://www.prnewswire.com/rss/technology-latest-news.rss", "press", 4),
        ("Business Wire", "https://businesswire.com", "https://feed.businesswire.com/rss/home/?rss=G1QFDERJXkJeGVJSVg==", "press", 4),
        ("GlobeNewsWire", "https://globenewswire.com", "https://www.globenewswire.com/RssFeed/subjectcode/26-Technology/feedTitle/GlobeNewswire - Technology", "press", 3),
    ]

    inserted = 0
    for name, url, rss_url, source_type, priority in sources:
        existing = db.execute(
            "SELECT id FROM sources WHERE name = ?", (name,)
        ).fetchone()
        if existing:
            continue

        source_id = str(uuid.uuid4())
        db.execute(
            """INSERT INTO sources (id, name, url, rss_feed_url, type, priority)
            VALUES (?, ?, ?, ?, ?, ?)""",
            (source_id, name, url, rss_url, source_type, priority)
        )
        inserted += 1

    db.commit()
    print(f"Seeded {inserted} sources")


if __name__ == "__main__":
    init_db()

    # Import from Monday.com Excel
    excel_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        "FOR_MERGING_Companies_1770749412 (1).xlsx"
    )
    if os.path.exists(excel_path):
        import_monday_excel(excel_path)
    else:
        print(f"Excel file not found at {excel_path}")
        print("Skipping Monday.com import")

    seed_sources()
    print("Seeding complete!")
